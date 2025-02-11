import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import json
import os
from datetime import datetime
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NumberFormatDescriptor

class QuotationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("上海伦伟-报价单系统-Sprit.Zeng V3.0-测试版")
        self.root.geometry("1200x1000")

        # 设置全局字体（Windows兼容）
        self.font_style = ("微软雅黑", 10)  # 主字体
        self.bold_font = ("微软雅黑", 10, "bold")  # 加粗字体
        self.info_font = ("微软雅黑", 10)  # 信息框字体

        # 应用字体设置
        self.root.option_add("*Font", self.font_style)
        ttk.Style().configure("Treeview", font=("微软雅黑", 10))  # 表格字体

        # 绑定窗口最小化和恢复事件
        self.root.bind("<Unmap>", self.on_minimize)  # 窗口最小化时触发
        self.root.bind("<Map>", self.on_restore)     # 窗口恢复时触发

        # 列对应关系
        self.COLUMN_MAPPING = {
            "物料编码": "物料编码",
            "物料名称": "物料名称",
            "规格型号": "规格型号",
            "数量": "数量",
            "含税单价": "含税单价"
        }

        # 存储完整的产品数据
        self.full_product_data = []

        # 历史记录文件路径
        self.history_file = self.get_history_file_path()

        # 初始化界面元素
        self.create_widgets()

        # 加载历史记录
        self.load_history_from_file()

    def on_minimize(self, event):
        """窗口最小化事件处理"""
        pass  # 这里可以添加窗口最小化时的处理逻辑，目前为空

    def on_restore(self, event):
        """窗口恢复事件处理"""
        self.root.deiconify()  # 强制显示窗口
        self.root.lift()       # 将窗口置于最上层
        self.root.focus_force()  # 强制聚焦窗口

    def get_history_file_path(self):
        """获取历史记录文件的路径，根据运行环境确定存储位置"""
        if getattr(sys, 'frozen', False):
            # 打包后的运行环境
            base_path = Path(sys._MEIPASS)
        else:
            # 开发环境
            base_path = Path(os.path.abspath("."))

        # 获取用户的应用程序数据目录，根据操作系统判断
        if sys.platform == "win32":
            app_data_dir = Path(os.getenv('APPDATA'))
        elif sys.platform == "darwin":
            app_data_dir = Path.home() / "Library" / "Application Support"
        else:
            app_data_dir = Path.home() / ".local" / "share"

        # 创建应用程序特定的子目录
        app_name = "QuotationApp"
        app_data_dir = app_data_dir / app_name
        app_data_dir.mkdir(parents=True, exist_ok=True)

        # 历史记录文件的路径
        history_file = app_data_dir / "quotation_history.json"

        # 如果文件不存在，则创建一个空文件
        if not history_file.exists():
            with open(history_file, "w", encoding="utf-8") as file:
                json.dump([], file)

        return str(history_file)

    def create_widgets(self):
        """创建所有UI界面元素"""
        # 主布局框架
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 工具栏框架
        toolbar = tk.Frame(main_frame)
        toolbar.pack(fill=tk.X, pady=5)

        # 保存按钮
        btn_save = tk.Button(toolbar, text="保存报价单", command=self.save_quotation)
        btn_save.pack(side=tk.RIGHT, padx=5)

        # 导出按钮
        btn_export = tk.Button(toolbar, text="导出报价单", command=self.export_excel)
        btn_export.pack(side=tk.RIGHT, padx=5)

        # 导入按钮 - 确保 command 参数正确指向 self.import_excel
        btn_import = tk.Button(toolbar, text="导入Excel", command=self.import_excel) # 确保这里是 self.import_excel
        btn_import.pack(side=tk.LEFT, padx=5)

        # 搜索框
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(toolbar, textvariable=self.search_var, width=40, font=self.font_style)
        self.search_entry.pack(side=tk.LEFT, padx=10)
        self.search_entry.bind("<KeyRelease>", self.filter_products)

        # 产品表格框架
        self.product_frame = tk.LabelFrame(main_frame, text="产品列表", font=self.bold_font)
        self.product_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.create_product_table() # 创建产品表格

        # 报价单表格框架
        self.quotation_frame = tk.LabelFrame(main_frame, text="报价单", font=self.bold_font)
        self.quotation_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.create_quotation_table() # 创建报价单表格

        # 底部信息框架
        bottom_frame = tk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=10)

        # 第一行：含税总价和毛利率框架
        row1_frame = tk.Frame(bottom_frame)
        row1_frame.pack(fill=tk.X, pady=5)

        # 含税总价标签和只读输入框
        tk.Label(row1_frame, text="含税成本总价：", font=self.font_style).pack(side=tk.LEFT, padx=(20, 0))
        self.total_label = ttk.Entry(row1_frame, font=self.bold_font, state="readonly", width=15)
        self.total_label.pack(side=tk.LEFT, padx=5)

        tk.Label(row1_frame, text="大写金额：", font=self.font_style).pack(side=tk.LEFT, padx=(20, 0))
        self.total_cn_label = ttk.Entry(row1_frame, font=self.bold_font, state="readonly", width=40)
        self.total_cn_label.pack(side=tk.LEFT, padx=5)

        # 毛利率输入框
        tk.Label(row1_frame, text="毛利率（%）：", font=self.font_style).pack(side=tk.LEFT, padx=(40, 0))
        self.profit_margin_entry = ttk.Entry(row1_frame, width=10, font=self.font_style)
        self.profit_margin_entry.pack(side=tk.LEFT, padx=5)
        self.profit_margin_entry.bind("<KeyRelease>", self.calculate_total) # 绑定键盘释放事件，实时计算总价

        # 第二行：最终含税总价框架
        row2_frame = tk.Frame(bottom_frame)
        row2_frame.pack(fill=tk.X, pady=5)

        tk.Label(row2_frame, text="最终含税总价：", font=self.font_style).pack(side=tk.LEFT, padx=(20, 0))
        self.final_total_label = ttk.Entry(row2_frame, font=self.bold_font, state="readonly", width=15)
        self.final_total_label.pack(side=tk.LEFT, padx=5)

        tk.Label(row2_frame, text="大写金额：", font=self.font_style).pack(side=tk.LEFT, padx=(20, 0))
        self.final_total_cn_label = ttk.Entry(row2_frame, font=self.bold_font, state="readonly", width=40)
        self.final_total_cn_label.pack(side=tk.LEFT, padx=5)

        # 清空配置按钮
        btn_clear = tk.Button(row2_frame, text="清空配置", font=self.font_style, command=self.clear_quotation)
        btn_clear.pack(side=tk.LEFT, padx=(30, 0))

        # 主内容区域框架
        content_frame = tk.Frame(main_frame, height=200) # 设置高度
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 左侧区域框架
        left_frame = tk.Frame(content_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 选中商品信息框架
        self.selected_item_frame = tk.LabelFrame(left_frame, text="选中商品信息", font=self.bold_font, height=50) # 设置高度
        self.selected_item_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # 选中商品信息文本框
        self.selected_item_info = tk.Text(self.selected_item_frame, font=self.info_font)
        self.selected_item_info.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 右侧历史记录区域框架
        history_frame = tk.LabelFrame(content_frame, text="历史报价记录", font=self.bold_font, height=50) # 设置高度
        history_frame.pack(side=tk.RIGHT, fill=tk.BOTH, padx=10, pady=10, ipadx=10, ipady=5)

        # 历史记录表格
        self.history_tree = ttk.Treeview(history_frame, columns=("时间", "总金额", "毛利率", "操作"), show="headings")
        self.history_tree.heading("时间", text="时间")
        self.history_tree.heading("总金额", text="总金额")
        self.history_tree.heading("毛利率", text="毛利率")
        self.history_tree.heading("操作", text="操作")
        self.history_tree.column("时间", width=150)
        self.history_tree.column("总金额", width=100)
        self.history_tree.column("毛利率", width=80)
        self.history_tree.column("操作", width=50, anchor="center")
        self.history_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 绑定历史记录删除事件
        self.history_tree.bind("<Button-1>", self.delete_history_item)

        # 历史记录操作按钮框架
        btn_frame = tk.Frame(history_frame)
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)

        # 删除记录按钮
        btn_delete = tk.Button(btn_frame, text="删除记录", font=self.font_style, command=self.delete_history)
        btn_delete.pack(side=tk.LEFT, padx=5)

        # 绑定历史记录双击事件，加载历史报价单
        self.history_tree.bind("<Double-1>", self.load_history_quotation)

    def create_product_table(self):
        """创建产品表格"""
        columns = list(self.COLUMN_MAPPING.values()) # 从COLUMN_MAPPING获取列名
        self.product_tree = ttk.Treeview(self.product_frame, columns=columns, show="headings") # 创建表格

        # 设置列宽
        col_widths = [150, 200, 250, 80, 100]
        for col, width in zip(columns, col_widths): # 遍历列名和宽度
            self.product_tree.heading(col, text=col) # 设置列标题
            self.product_tree.column(col, width=width, anchor="center") # 设置列宽和对齐方式

        # 添加滚动条
        scrollbar = ttk.Scrollbar(self.product_frame, orient=tk.VERTICAL, command=self.product_tree.yview) # 创建垂直滚动条
        self.product_tree.configure(yscroll=scrollbar.set) # 配置表格的y轴滚动
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y) # 滚动条靠右，垂直填充
        self.product_tree.pack(fill=tk.BOTH, expand=True) # 表格填充框架

        # 绑定双击事件，添加到报价单
        self.product_tree.bind("<Double-1>", self.add_to_quotation)

        # 绑定 Ctrl+C 复制事件
        self.product_tree.bind("<Control-c>", self.copy_selected_text)

    def copy_selected_text(self, event):
        """复制产品表格中选中的内容到剪贴板"""
        selected_items = self.product_tree.selection() # 获取选中行
        if not selected_items: # 如果没有选中行，则返回
            return

        # 获取选中行的文本内容
        all_text = ""
        for item_id in selected_items: # 遍历选中的每一行
            item_values = self.product_tree.item(item_id, 'values') # 获取行数据
            text_line = "\t".join(map(str, item_values)) + "\n" # 将行数据转换为制表符分隔的字符串，列之间用制表符分隔，行尾换行
            all_text += text_line # 添加到总文本

        # 复制到剪贴板
        self.root.clipboard_clear() # 清空剪贴板
        self.root.clipboard_append(all_text) # 将文本添加到剪贴板
        self.root.update() # 更新剪贴板，立即生效


    def create_quotation_table(self):
        """创建报价单表格"""
        columns = ("物料编码", "物料名称", "规格型号", "数量", "含税单价", "小计", "操作") # 定义报价单表格列名
        self.quotation_tree = ttk.Treeview(self.quotation_frame, columns=columns, show="headings") # 创建报价单表格

        # 设置列宽
        col_widths = [120, 150, 350, 80, 80, 80, 60]
        for col, width in zip(columns, col_widths): # 遍历列名和宽度
            self.quotation_tree.heading(col, text=col) # 设置列标题
            self.quotation_tree.column(col, width=width, anchor="center") # 设置列宽和对齐方式

        # 添加滚动条
        scrollbar = ttk.Scrollbar(self.quotation_frame, orient=tk.VERTICAL, command=self.quotation_tree.yview) # 创建垂直滚动条
        self.quotation_tree.configure(yscroll=scrollbar.set) # 配置表格y轴滚动
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y) # 滚动条靠右，垂直填充
        self.quotation_tree.pack(fill=tk.BOTH, expand=True) # 表格填充框架

        # 绑定删除事件（Delete键）
        self.quotation_tree.bind("<Delete>", self.delete_item)

        # 绑定选中事件，显示商品信息
        self.quotation_tree.bind("<<TreeviewSelect>>", self.show_selected_item_info)

        # 绑定双击事件，编辑数量和单价
        self.quotation_tree.bind("<Double-1>", self.edit_quotation_item)

        # 绑定操作列的点击事件，删除行
        self.quotation_tree.bind("<Button-1>", self.handle_operation_click)

        # 用于存储编辑框
        self.quotation_edit_entry = None
        self.quotation_edit_column = None
        self.quotation_edit_item = None

    def handle_operation_click(self, event):  # 修复 handle_operation_click 的 self 缺失  !!! 移除重复的 self !!!
        """处理操作列的点击事件，用于删除报价单行"""
        region = self.quotation_tree.identify_region(event.x, event.y) # 识别点击区域
        if region != "cell": # 如果不是单元格区域，则返回
            return

        column = self.quotation_tree.identify_column(event.x) # 识别点击列
        item = self.quotation_tree.identify_row(event.y) # 识别点击行

        # 只处理操作列（第7列）
        if column == "#7": # 操作列索引为 #7
            self.quotation_tree.delete(item) # 删除选中行
            self.calculate_total()  # 重新计算总价

    def edit_quotation_item(self, event):
        """双击报价单中的数量或含税单价列，进入编辑模式"""
        region = self.quotation_tree.identify_region(event.x, event.y) # 识别点击区域
        if region != "cell": # 如果不是单元格区域，则返回
            return

        column = self.quotation_tree.identify_column(event.x) # 识别点击列
        item = self.quotation_tree.identify_row(event.y) # 识别点击行

        # 只允许编辑“数量”和“含税单价”列
        if column not in ("#4", "#5"): # 数量列索引为 #4，含税单价列索引为 #5
            return

        # 获取当前单元格的值
        values = self.quotation_tree.item(item, "values") # 获取行数据
        col_index = int(column[1:]) - 1  # 列索引从0开始，例如 #4 对应索引 3
        cell_value = values[col_index] # 获取单元格值

        # 获取单元格的坐标和大小
        x, y, width, height = self.quotation_tree.bbox(item, column) # 获取单元格位置和大小

        # 创建编辑框
        self.quotation_edit_entry = ttk.Entry(self.quotation_frame, width=width, font=self.font_style) # 创建Entry编辑框
        self.quotation_edit_entry.place(x=x, y=y, width=width, height=height) # 放置编辑框在单元格位置
        self.quotation_edit_entry.insert(0, cell_value) # 初始值设为单元格当前值
        self.quotation_edit_entry.focus() # 聚焦到编辑框

        # 绑定回车键和失去焦点事件
        self.quotation_edit_entry.bind("<Return>", lambda e: self.save_quotation_edit(item, col_index)) # 回车保存
        self.quotation_edit_entry.bind("<FocusOut>", lambda e: self.save_quotation_edit(item, col_index)) # 失去焦点保存

        # 存储当前编辑的列和行
        self.quotation_edit_column = col_index
        self.quotation_edit_item = item

    def save_quotation_edit(self, item, col_index):
        """保存报价单编辑后的值，并更新小计和总价"""
        if self.quotation_edit_entry: # 确保编辑框存在
            new_value = self.quotation_edit_entry.get() # 获取编辑框的值

            # 更新报价单表格中的值
            values = list(self.quotation_tree.item(item, "values")) # 获取行数据并转换为列表
            values[col_index] = new_value # 更新指定列的值

            # 如果是数量或含税单价列，重新计算小计
            if col_index == 3:  # 数量列
                try:
                    quantity = float(new_value) # 转换为浮点数
                    unit_price = float(values[4]) # 获取含税单价
                    values[5] = f"{quantity * unit_price:.2f}"  # 更新小计，保留两位小数
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的数字！") # 错误提示
                    return # 停止保存
            elif col_index == 4:  # 含税单价列
                try:
                    unit_price = float(new_value) # 转换为浮点数
                    quantity = float(values[3]) # 获取数量
                    values[5] = f"{quantity * unit_price:.2f}"  # 更新小计，保留两位小数
                    values[4] = f"{unit_price:.2f}"  # 确保含税单价保留两位小数
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的数字！") # 错误提示
                    return # 停止保存

            # 更新行数据
            self.quotation_tree.item(item, values=values) # 更新表格行数据

            # 销毁编辑框
            self.quotation_edit_entry.destroy() # 销毁编辑框
            self.quotation_edit_entry = None # 清空编辑框变量

            # 重新计算总价
            self.calculate_total() # 更新总价

    def show_selected_item_info(self, event):
        """显示报价单中选中商品的详细信息在文本框中"""
        selected_item = self.quotation_tree.selection() # 获取选中的行
        if not selected_item: # 如果没有选中行，则返回
            return

        # 获取选中行的数据
        values = self.quotation_tree.item(selected_item, "values") # 获取行数据

        # 清空原有内容
        self.selected_item_info.delete(1.0, tk.END) # 清空文本框

        # 显示详细信息
        info = (
            f"物料编码: {values[0]}\n"
            f"物料名称: {values[1]}\n"
            f"规格型号: {values[2]}\n"
            f"数量: {values[3]}\n"
            f"含税单价: {values[4]}\n"
            f"小计: {values[5]}\n"
        ) # 格式化商品信息
        self.selected_item_info.insert(tk.END, info) # 将信息插入文本框

    def save_quotation(self):
        """保存当前报价单到历史记录"""
        # 获取当前报价单数据
        quotation_data = []
        for item in self.quotation_tree.get_children(): # 遍历报价单表格的所有行
            values = self.quotation_tree.item(item, "values") # 获取每行的数据
            quotation_data.append({ # 将每行数据以字典形式添加到列表
                "物料编码": values[0],
                "物料名称": values[1],
                "规格型号": values[2],
                "数量": values[3],
                "含税单价": values[4],
                "小计": values[5]
            })

        # 获取当前时间和毛利率
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S") # 获取当前时间
        try:
            profit_margin = float(self.profit_margin_entry.get()) # 获取毛利率
        except ValueError:
            profit_margin = 0.0 # 默认毛利率为0

        # 获取总金额
        total_amount = self.total_label.get().replace(",", "") # 获取总金额，并移除千分位分隔符

        # 添加到历史记录表格
        self.history_tree.insert("", "end", values=(current_time, total_amount, f"{profit_margin:.2f}%", "删除")) # 添加历史记录到表格

        # 保存到文件
        self.save_history_to_file(current_time, total_amount, profit_margin, quotation_data) # 调用函数保存到文件

    def save_history_to_file(self, current_time, total_amount, profit_margin, quotation_data):
        """将历史记录保存到JSON文件"""
        history_entry = {
            "时间": current_time,
            "总金额": total_amount,
            "毛利率": f"{profit_margin:.2f}%",
            "报价单详情": quotation_data  # 添加报价单详细数据
        } # 创建历史记录条目，包含时间、总金额、毛利率和报价单详情

        # 读取现有历史记录
        try:
            with open(self.history_file, "r", encoding="utf-8") as file: # 尝试读取历史记录文件
                history = json.load(file) # 从JSON文件加载数据
        except FileNotFoundError:
            history = [] # 文件不存在则初始化为空列表

        # 添加新记录
        history.append(history_entry) # 将新的历史记录条目添加到历史记录列表

        # 保存更新后的历史记录
        with open(self.history_file, "w", encoding="utf-8") as file: # 以写入模式打开历史记录文件
            json.dump(history, file, ensure_ascii=False, indent=4) # 将更新后的历史记录写入JSON文件，并格式化

    def load_history_from_file(self):
        """从JSON文件加载历史记录并在历史记录表格中显示"""
        if not os.path.exists(self.history_file): # 如果历史记录文件不存在，则直接返回
            return

        with open(self.history_file, "r", encoding="utf-8") as f: # 以只读模式打开历史记录文件
            history_data = json.load(f) # 从JSON文件加载历史记录数据

        for entry in history_data: # 遍历历史记录数据
            self.history_tree.insert("", "end", values=(entry["时间"], entry["总金额"], entry["毛利率"], "删除")) # 将每条历史记录添加到历史记录表格

    def export_excel(self):
        """导出报价单到Excel文件，使用模板，动态调整行数，自动序号，数值转换, 格式复制"""
        col_mapping = { # 列位置映射关系，定义在函数最前面
            "序号": 2,
            "物料名称": 3,
            "规格型号": 4,
            "数量": 5,
            "含税单价": 6,
            "小计": 7
        }
        try:
            # 读取模板文件
            template_path = "Quotation.xlsx"
            wb = load_workbook(template_path)
            ws = wb.active

            # 获取当前报价单数据
            quotation_data = []
            for item in self.quotation_tree.get_children():
                values = self.quotation_tree.item(item, "values")
                quotation_data.append({
                    "物料名称": values[1],
                    "规格型号": values[2],
                    "数量": values[3],
                    "含税单价": values[4],
                    "小计": values[5]
                })

            start_row = 19
            end_row = 28
            template_product_rows = end_row - start_row + 1
            quotation_item_count = len(quotation_data)

            # 动态调整行数
            if quotation_item_count > template_product_rows:
                rows_to_insert = quotation_item_count - template_product_rows
                insert_row_index = end_row + 1
                ws.insert_rows(idx=insert_row_index, amount=rows_to_insert)


            elif quotation_item_count < template_product_rows:
                # 不删除行，而是清除多余行的内容
                start_clear_row = start_row + quotation_item_count
                end_clear_row = start_row + template_product_rows - 1
                for row_num in range(start_clear_row, end_clear_row + 1):
                    for col_name in col_mapping.keys():
                        col_idx = col_mapping[col_name]
                        cell = ws.cell(row=row_num, column=col_idx)
                        cell.value = None  # 清空单元格内容


            # 清空并填充产品数据区域，并设置对齐方式和自动换行
            current_end_row = start_row + quotation_item_count - 1
            for i, item in enumerate(quotation_data):
                current_row = start_row + i
                # 自动填充序号
                cell_no = ws.cell(row=current_row, column=col_mapping["序号"])
                cell_no.value = i + 1
                cell_no.alignment = Alignment(horizontal='center', vertical='center') # 序号居中对齐

                # 填充物料名称
                cell_name = ws.cell(row=current_row, column=col_mapping["物料名称"])
                cell_name.value = item["物料名称"]
                cell_name.alignment = Alignment(horizontal='center', vertical='center') # 物料名称居中对齐

                # 填充规格型号
                cell_spec = ws.cell(row=current_row, column=col_mapping["规格型号"])
                cell_spec.value = item["规格型号"]
                cell_spec.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True) # 规格型号左对齐, 自动换行

                # 填充数量
                cell_qty = ws.cell(row=current_row, column=col_mapping["数量"])
                cell_qty.value = float(item["数量"]) # 数量转换为数值类型
                cell_qty.alignment = Alignment(horizontal='center', vertical='center') # 数量居中对齐

                # 填充含税单价
                cell_price = ws.cell(row=current_row, column=col_mapping["含税单价"])
                cell_price.value = float(item["含税单价"].replace(",", "")) # 移除千分位分隔符, 转换为数值
                cell_price.alignment = Alignment(horizontal='right', vertical='center') # 含税单价右对齐

                # 填充小计
                cell_subtotal = ws.cell(row=current_row, column=col_mapping["小计"])
                cell_subtotal.value = float(item["小计"].replace(",", "")) # 移除千分位分隔符, 转换为数值
                cell_subtotal.alignment = Alignment(horizontal='right', vertical='center') # 小计右对齐


            # 计算总价并填充 (应用 Total 行样式)
            total_amount = self.total_label.get().replace(",", "")
            total_cell = ws.cell(row=29, column=col_mapping["小计"]) # **!!!  Total 行 行号再次确认为 29 !!!**
            if total_cell.coordinate not in ws.merged_cells.ranges:
                try:
                    total_amount_float = float(total_amount)
                except:
                    total_amount_float = 0.0
                total_cell.value = total_amount_float


            # 弹出保存文件对话框 (保持不变)
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("导出成功", f"报价单已成功导出到 {file_path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出Excel文件时出错：{e}")


    def import_excel(self):
        """导入Excel产品数据到产品表格"""
        file_path = filedialog.askopenfilename( # 弹出文件选择对话框，让用户选择Excel文件
            title="选择 Excel 文件", # 对话框标题
            filetypes=[("Excel files", "*.xlsx *.xls")] # 文件类型过滤器，只显示Excel文件
        )

        if file_path: # 如果用户选择了文件
            success, message = self.load_excel_data(file_path) # 调用 load_excel_data 函数加载Excel数据
            if success: # 如果加载成功
                messagebox.showinfo("导入成功", message) # 弹出导入成功提示框
            else: # 如果加载失败
                messagebox.showerror("导入失败", message) # 弹出导入失败错误提示框

    def load_excel_data(self, file_path):
        """加载Excel数据到产品表格，并处理数据格式"""
        try:
            # 强制将“规格型号”列读取为字符串类型，避免可能的类型推断错误
            df = pd.read_excel(file_path, dtype={"规格型号": str}) # 读取Excel文件到DataFrame

            # 检查Excel文件是否包含必需的列
            required_columns = list(self.COLUMN_MAPPING.keys()) # 获取必需的列名列表
            if not all(col in df.columns for col in required_columns): # 检查所有必需列是否都存在于DataFrame中
                return False, f"Excel 文件缺少必需的列：{required_columns}" # 如果缺少列，返回False和错误信息

            # 清空现有产品表格
            self.product_tree.delete(*self.product_tree.get_children()) # 删除产品表格的所有行

            # 清空完整产品数据列表
            self.full_product_data = []  # 清空完整数据，准备加载新数据

            # 遍历DataFrame的每一行，将数据加载到产品表格中
            for index, row in df.iterrows(): # 遍历DataFrame的行
                values = [row[col] for col in required_columns] # 提取每行中必需列的值，按COLUMN_MAPPING顺序
                # 确保含税单价保留两位小数，并转换为字符串类型
                values[4] = f"{float(values[4]):.2f}" # 格式化含税单价，保留两位小数
                self.product_tree.insert("", "end", values=values) # 将数据插入产品表格末尾
                self.full_product_data.append(values)  # 将数据添加到完整产品数据列表

            return True, f"成功导入 {len(df)} 条产品数据！" # 返回True和成功导入的消息
        except Exception as e: # 捕获加载Excel数据过程中的异常
            return False, f"导入 Excel 文件时出错：{e}" # 返回False和错误信息

    def filter_products(self, event=None):
        """根据搜索框内容过滤产品表格"""
        search_term = self.search_var.get().strip().lower()  # 获取搜索框内容，去除首尾空格并转换为小写

        # 清空当前显示的产品列表
        self.product_tree.delete(*self.product_tree.get_children()) # 删除产品表格的所有行，准备重新加载过滤后的数据

        # 遍历完整的产品数据，进行过滤
        for values in self.full_product_data: # 遍历完整产品数据列表
            spec = str(values[2]).lower()  # 获取“规格型号”列的值，并转换为小写

            # 如果搜索内容为空 或 “规格型号”包含搜索内容，则显示该行
            if not search_term or search_term in spec: # 检查是否满足过滤条件
                self.product_tree.insert("", "end", values=values) # 将满足条件的产品数据插入产品表格

    def add_to_quotation(self, event):
        """双击产品列表中的产品，将其添加到报价单表格"""
        selected_item = self.product_tree.selection() # 获取产品表格中选中的行
        if not selected_item: # 如果没有选中任何行，则返回
            return # 添加 return statement here

        # 获取选中行的数据
        item_values = self.product_tree.item(selected_item, "values") # 获取选中行的数据
        material_code = item_values[0]  # 物料编码
        material_name = item_values[1]  # 物料名称
        spec = item_values[2]  # 规格型号
        quantity = 1  # 默认数量为 1
        unit_price = float(item_values[4])  # 含税单价

        # 检查报价单表格中是否已存在相同物料编码的产品
        existing_item = None
        for item in self.quotation_tree.get_children(): # 遍历报价单表格的所有行
            values = self.quotation_tree.item(item, "values") # 获取每行的数据
            if values[0] == material_code:  # 通过物料编码判断是否已存在相同产品
                existing_item = item # 如果已存在，则记录该行
                break # 找到已存在的行后，跳出循环

        if existing_item: # 如果报价单中已存在相同产品
            # 如果已存在，则增加数量
            existing_values = self.quotation_tree.item(existing_item, "values") # 获取已存在行的数据
            new_quantity = int(existing_values[3]) + quantity  # 新数量 = 原有数量 + 默认数量
            subtotal = new_quantity * unit_price  # 重新计算小计
            self.quotation_tree.item(existing_item, values=( # 更新已存在行的数据
                material_code, material_name, spec, new_quantity, f"{unit_price:.2f}", f"{subtotal:.2f}", "删除"
            ))
        else: # 如果报价单中不存在相同产品
            # 如果不存在，则新增一行
            subtotal = quantity * unit_price  # 计算小计
            self.quotation_tree.insert("", "end", values=( # 在报价单表格末尾插入新行
                material_code, material_name, spec, quantity, f"{unit_price:.2f}", f"{subtotal:.2f}", "删除"
            ))

        # 更新总价
        self.calculate_total() # 重新计算报价单总价

    def calculate_total(self, event=None):
        """计算报价表的总价，包括含税总价、大写金额和最终含税总价"""
        total = 0.0 # 初始化总价
        for item in self.quotation_tree.get_children(): # 遍历报价单表格的所有行
            values = self.quotation_tree.item(item, "values") # 获取每行的数据
            subtotal = float(values[5])  # 获取小计，并转换为浮点数
            total += subtotal # 累加小计到总价

        # 更新含税总价（显示千分位分隔符）
        self.total_label.config(state="normal") # 设置为可编辑状态
        self.total_label.delete(0, tk.END) # 清空原有内容
        self.total_label.insert(0, f"{total:,.2f}")  # 格式化总价，保留两位小数，并使用千分位分隔符
        self.total_label.config(state="readonly") # 设置回只读状态

        # 更新大写金额
        self.total_cn_label.config(state="normal") # 设置为可编辑状态
        self.total_cn_label.delete(0, tk.END) # 清空原有内容
        self.total_cn_label.insert(0, self.to_chinese_amount(total)) # 将总价转换为大写金额并显示
        self.total_cn_label.config(state="readonly") # 设置回只读状态

        # 计算最终含税总价（考虑毛利率）
        try:
            profit_margin = float(self.profit_margin_entry.get()) # 获取毛利率
            final_total = total * (1 + profit_margin / 100) # 计算最终总价
            self.final_total_label.config(state="normal") # 设置为可编辑状态
            self.final_total_label.delete(0, tk.END) # 清空原有内容
            self.final_total_label.insert(0, f"{final_total:,.2f}")  # 格式化最终总价，保留两位小数，并使用千分位分隔符
            self.final_total_label.config(state="readonly") # 设置回只读状态

            # 更新最终大写金额
            self.final_total_cn_label.config(state="normal") # 设置为可编辑状态
            self.final_total_cn_label.delete(0, tk.END) # 清空原有内容
            self.final_total_cn_label.insert(0, self.to_chinese_amount(final_total)) # 将最终总价转换为大写金额并显示
            self.final_total_cn_label.config(state="readonly") # 设置回只读状态
        except ValueError:
            pass # 如果毛利率输入框内容无法转换为数字，则忽略，不计算最终总价

    def delete_item(self, event):
        """删除报价单中选中的行"""
        selected_item = self.quotation_tree.selection() # 获取报价单表格中选中的行
        if selected_item: # 如果有选中行
            self.quotation_tree.delete(selected_item) # 删除选中行
            self.calculate_total()  # 重新计算总价

    def clear_quotation(self):
        """清空报价单表格"""
        for item in self.quotation_tree.get_children(): # 遍历报价单表格的所有行
            self.quotation_tree.delete(item) # 删除每一行
        self.calculate_total()  # 重置总价

    def delete_history_item(self, event):
        """删除历史记录中的某一行，并同步更新历史记录文件"""
        region = self.history_tree.identify_region(event.x, event.y) # 识别点击区域
        if region != "cell": # 如果不是单元格区域，则返回
            return

        column = self.history_tree.identify_column(event.x) # 识别点击列
        item = self.history_tree.identify_row(event.y) # 识别点击行

        # 只处理操作列（第4列）
        if column == "#4": # 操作列索引为 #4
            # 获取选中行的时间戳
            time_str = self.history_tree.item(item, "values")[0] # 获取选中行的第一列值，即时间戳

            # 从文件中删除对应的历史记录
            with open(self.history_file, "r", encoding="utf-8") as file: # 以只读模式打开历史记录文件
                history = json.load(file) # 从JSON文件加载数据

            # 找到并删除对应的记录
            history = [entry for entry in history if entry["时间"] != time_str] # 过滤掉要删除的记录

            # 保存更新后的历史记录到文件
            with open(self.history_file, "w", encoding="utf-8") as file: # 以写入模式打开历史记录文件
                json.dump(history, file, ensure_ascii=False, indent=4) # 将更新后的历史记录写入JSON文件，并格式化

            # 从界面中删除该行
            self.history_tree.delete(item) # 从历史记录表格中删除选中行

    def delete_history(self):
        """删除所有历史记录，并同步更新历史记录文件"""
        # 清空文件内容
        with open(self.history_file, "w", encoding="utf-8") as file: # 以写入模式打开历史记录文件
            json.dump([], file, ensure_ascii=False, indent=4) # 将空列表写入JSON文件，清空文件内容

        # 清空界面中的历史记录表格
        for item in self.history_tree.get_children(): # 遍历历史记录表格的所有行
            self.history_tree.delete(item)

    def load_history_quotation(self, event):
        """双击历史记录加载报价单数据"""
        selected_item = self.history_tree.selection() # 获取选中的历史记录行
        if not selected_item: # 如果没有选中任何记录，则返回
            return

        history_values = self.history_tree.item(selected_item, "values") # 获取选中历史记录行的数据
        time_str = history_values[0] # 获取时间戳，用于查找对应的报价单详情

        # 从历史记录文件中查找对应的报价单详情
        with open(self.history_file, "r", encoding="utf-8") as file: # 以只读模式打开历史记录文件
            history_data = json.load(file) # 加载历史记录数据

        quotation_detail = None # 初始化报价单详情为 None
        for entry in history_data: # 遍历历史记录条目
            if entry["时间"] == time_str: # 找到匹配时间戳的记录
                quotation_detail = entry["报价单详情"] # 获取报价单详情
                profit_margin_str = entry.get("毛利率", "0.0%") # 获取毛利率，如果不存在则默认为 "0.0%"
                profit_margin = float(profit_margin_str.replace("%", "")) # 移除百分号并转换为浮点数
                break # 找到后跳出循环

        if quotation_detail: # 如果找到了报价单详情
            # 清空当前报价单表格
            for item in self.quotation_tree.get_children(): # 遍历并删除报价单表格所有行
                self.quotation_tree.delete(item)

            # 将历史报价单数据填充到报价单表格
            for item_data in quotation_detail: # 遍历报价单详情数据
                self.quotation_tree.insert("", "end", values=( # 插入新行
                    item_data["物料编码"],
                    item_data["物料名称"],
                    item_data["规格型号"],
                    item_data["数量"],
                    item_data["含税单价"],
                    item_data["小计"],
                    "删除"  # 操作列
                ))
            # 将毛利率数据填充到毛利率输入框
            self.profit_margin_entry.delete(0, tk.END) # 清空毛利率输入框
            self.profit_margin_entry.insert(0, f"{profit_margin:.2f}") # 填充毛利率

            # 重新计算总价
            self.calculate_total() # 重新计算总价


    def to_chinese_amount(self, amount):
        """
        将数字金额转换为中文大写金额。

        Args:
            amount (float or int): 数字金额。

        Returns:
            str: 中文大写金额字符串。
        """
        if not isinstance(amount, (int, float)):
            return "金额类型错误"
        if amount < 0:
            return "负数金额不支持"
        if amount == 0:
            return "零元"

        integer_part = int(amount)
        decimal_part = round((amount - integer_part) * 100)

        CN_NUM = ["零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖"]
        CN_UNIT = ["元", "拾", "佰", "仟", "万", "亿"]
        CN_DECIMAL_UNIT = ["", "角", "分"]

        def integer_to_chinese(num):
            if num == 0:
                return ""
            chinese_integer = ""
            unit_index = 0
            while num > 0:
                digit = num % 10
                if digit != 0:
                    chinese_integer = CN_NUM[digit] + CN_UNIT[unit_index] + chinese_integer
                else:
                    if unit_index % 4 == 0 and chinese_integer:
                        chinese_integer = CN_UNIT[unit_index] + chinese_integer
                    elif unit_index % 4 != 0 and chinese_integer and chinese_integer[0] != "零":
                        chinese_integer = "零" + chinese_integer
                num //= 10
                unit_index += 1
            return chinese_integer

        chinese_amount = integer_to_chinese(integer_part)
        if decimal_part > 0:
            yuan_suffix = "元" if chinese_amount else "元" # 金额为零元也要有元字
            chinese_amount += yuan_suffix
            if decimal_part < 10:
                chinese_amount += "零" + CN_NUM[decimal_part // 10] + CN_DECIMAL_UNIT[2] # 0X分
            else:
                chinese_amount += CN_NUM[decimal_part // 10] + CN_DECIMAL_UNIT[1] # X角
                if decimal_part % 10 != 0:
                     chinese_amount += CN_NUM[decimal_part % 10] + CN_DECIMAL_UNIT[2] # X分
        else:
             chinese_amount += "元整" #  整数金额以元整结尾

        # 移除多余的零，并处理 Yuan 和 万 亿 单位
        chinese_amount = chinese_amount.replace("零元", "元").replace("零万", "万").replace("零亿", "亿").replace("零零", "零").rstrip("零").replace("元万", "万").replace("元亿", "亿").strip("零")
        if chinese_amount.startswith("元"): # 如果以元开头，则移除元字
            chinese_amount = chinese_amount[1:]
        if not chinese_amount:
            return "零元" # 如果为空，返回零元

        return chinese_amount

if __name__ == "__main__":
    root = tk.Tk()
    app = QuotationApp(root)
    root.mainloop()
